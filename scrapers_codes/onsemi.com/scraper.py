import logging
import time
import sys
import traceback
import os
import openpyxl
from openpyxl.styles import Font
import argparse
from selenium import webdriver
from selenium.common import exceptions
from datetime import datetime
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
import re

logging.basicConfig(format='%(levelname)s %(asctime)s:  %(message)s', level=logging.INFO)
from selenium.webdriver.remote.remote_connection import LOGGER
from urllib3.connectionpool import log as urllibLogger

urllibLogger.setLevel(logging.WARNING)
LOGGER.setLevel(logging.WARNING)
root_dir = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
error_reports_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), "error_reports")
os.makedirs(error_reports_dir, exist_ok=True)
output_dir = os.path.join(root_dir, "results_data_frames", "onsemi.com")
parser = argparse.ArgumentParser(
    description="This program will get products from onseim.com and put results to output file.")
parser.add_argument('--output-file-path', '-o', action="store", help="path for output json file.",
                    default=os.path.join(output_dir, 'products.xlsx'))
parser.add_argument('--browser', action="store_true", help="Turn off headless mode to see browser.")
parser.add_argument('--skip-products', type=int, default=0, action="store", help="skip products.")
args = parser.parse_args()


class Scraper:
    def __init__(self):
        self.max_tries = 3
        self.wait_time = 0.5
        self.results = []
        self.row_count = 1
        self.output_path = args.output_file_path
        self.skip_products = args.skip_products
        os.makedirs(output_dir, exist_ok=True)
        logging.info(f"Output path set: {self.output_path}")
        # self.ofh = open(output_path, 'w')
        if self.skip_products > 0:
            if os.path.isfile(self.output_path):
                self.wb = openpyxl.load_workbook(self.output_path)
            else:
                raise Exception(
                    "When skip-products value given, previous excel file must be there on output path to append data.")
        else:
            self.wb = openpyxl.Workbook()
        # self.sheet = self.wb.create_sheet()
        self.sheet = self.wb.active
        self.options = webdriver.ChromeOptions()
        if not args.browser:
            self.options.add_argument("--headless")
            self.options.add_argument("--start-maximized")
        self.options.add_argument("--log-level=3")
        self.cd = webdriver.Chrome(options=self.options, service_log_path='NUL')

        is_driver_quit = False
        try:
            self.start_job()
        except KeyboardInterrupt:
            logging.warning("Keyboard Interrupt. Closing...")
            is_driver_quit = True
        except:
            try:
                filename = os.path.join(error_reports_dir,
                                        f"exception_{datetime.now().strftime('%Y-%m-%dT%I-%M-%S-%p')}")
                efh = open(filename, 'w')
                traceback.print_exc(file=efh)
                efh.close()
                self.cd.save_screenshot(f'{filename}.png')
                hfh = open(filename + ".html", 'w', encoding='utf-8')
                hfh.write(str(self.cd.page_source))
                hfh.close()
                logging.info(f"Got some exception. Quitting the program. See details in file: {filename}.")
            except:
                traceback.print_exc(file=sys.stdout)
        finally:
            logging.info(f"Total Products found: {len(self.results)}.")
            self.wb.save(self.output_path)
            # self.ofh.write(json.dumps(self.results, indent=4))
            # self.ofh.close()
            logging.info(f"Output stored to {self.output_path}")
            if not is_driver_quit:
                self.cd.quit()

    def add_row_to_sheet(self, row, bold=False, row_inc=1):
        for c, item in enumerate(row):
            v = re.sub(ILLEGAL_CHARACTERS_RE, "?", item)
            self.sheet.cell(self.row_count, c + 1, value=v).font = Font(bold=bold)
        self.row_count += row_inc

    def start_job(self):
        logging.info("Job started. Finding product detail pages...")
        self.cd.get("https://www.onsemi.com/products")
        anchors_without_sub_cat = self.cd.find_elements_by_xpath(
            "//div[not(@data-toggle) and h4]/following-sibling::div//h6[not(./following-sibling::div)]/a")
        sub_cat_anchors = self.cd.find_elements_by_xpath(
            "//div[not(@data-toggle) and h4]/following-sibling::div//h6/following-sibling::div/a")

        detail_pages_urls = [e.get_attribute('href') for e in anchors_without_sub_cat] + \
                            [e.get_attribute('href') for e in sub_cat_anchors]
        logging.info(f'Total product details pages: {len(detail_pages_urls)}.')
        i, products_count = 0, 0
        while i < len(detail_pages_urls):
            url = detail_pages_urls[i]
            # j += 1
            i += 1
            logging.info(f'Working on page {i}, url = {url}')
            self.cd.get(url)
            self.cd.implicitly_wait(10)
            page_error_count = 0
            is_continue_page_loop = False
            while True:
                try:
                    self.cd.find_element_by_xpath("//select[@name='pageSize']/option[.='ALL']").click()
                    break
                except exceptions.NoSuchElementException:
                    new_urls_elements = self.cd.find_elements_by_xpath("//a[.='View Products']")
                    if len(new_urls_elements):
                        for new_url in new_urls_elements:
                            if new_url not in detail_pages_urls:
                                detail_pages_urls.append(new_url)
                        # detail_pages_urls += [new_url := e.get_attribute('href') for e in new_urls_elements if
                        #                       new_url not in detail_pages_urls]
                        logging.info(f"More product details pages URLs found on this page."
                                     f" Total product pages now: {len(detail_pages_urls)}")
                        is_continue_page_loop = True
                        break
                    else:
                        if page_error_count < 10:
                            # self.cd.execute_script(
                            #     "return document.body.innerText.includes('An error has occurred in our System. The erro"
                            #     "r message has been captured and will be investigated')") and
                            logging.error(f"Getting error on page {i}, Trying again ({page_error_count})")
                            self.cd.refresh()
                            page_error_count += 1
                        else:
                            raise
            if is_continue_page_loop:
                continue
            self.cd.implicitly_wait(0)
            time.sleep(2)
            retry_loading_count = 0
            while self.cd.find_elements_by_xpath(
                    "//div[@class='px-overlay']//div[@class='spinner-border green']"):
                if retry_loading_count > 10:
                    i -= 1
                    logging.info('Looks like page is stuck.')
                    continue
                logging.debug("New data still loading.")
                time.sleep(1)
            # self.cd.implicitly_wait(10)
            heading = self.cd.find_element_by_xpath("//div[@id='breadcrumb']/span").text
            self.add_row_to_sheet([heading], bold=True)
            field_names = [self.get_txt_by_xpath('.', e) for e in self.cd.find_elements_by_xpath(
                "//div[contains(@class, 'px-header-cell-heading')]")]
            _field_names = field_names.copy()
            _field_names.append('verify_url')
            _field_names.remove('Select') if 'Select' in field_names else None
            _field_names.remove("Data Sheet") if "Data Sheet" in field_names else None
            self.add_row_to_sheet(_field_names, bold=True)
            product_rows = self.cd.find_elements_by_xpath("//div[contains(@class, 'px-row ') and contains(@id, 'r_')]")
            logging.info(f"Products found: {len(product_rows)}")
            for row in product_rows:
                if self.skip_products > 0:
                    self.skip_products -= 1
                    self.row_count += 1
                    logging.info(f"skipping product. skip left {self.skip_products}")
                    products_count += 1
                    continue
                else:
                    page_data = {}
                    for fn in field_names:
                        page_data[fn] = self.get_txt_by_xpath(
                            f"./div[contains(@class, 'px-cell')][{field_names.index(fn) + 1}]"
                            , row)
                        logging.debug(f"{fn} = {page_data[fn]}")
                    page_data['verify_url'] = url
                    page_data.pop('Select', None)
                    page_data.pop("Data Sheet", None)
                    self.results.append(page_data)
                    self.add_row_to_sheet(page_data.values())
                products_count += 1
                logging.info(f"Total Products done: {products_count}")
                if products_count % 10 == 0:
                    self.wb.save(self.output_path)
            self.row_count += 1

    def click_by_xpath(self, xpath, element=None):
        result = False
        for i in range(self.max_tries):
            result = self.cd.execute_script(f"""
           //console.log(arguments);
            parent = document;
            if(arguments[1]){{
                parent = arguments[1];
            }}
            node = document.evaluate(arguments[0], parent, null,
            XPathResult.FIRST_ORDERED_NODE_TYPE, null ).singleNodeValue;
            //console.log(node);
            if(node){{
             node.scrollIntoView({{
            behavior: 'auto',
            block: 'center',
            inline: 'center'
            }});
             node.click(); 
             return true;
            }}
            return false;
                    """, xpath, element)
            if not result:
                time.sleep(self.wait_time)
            else:
                break
        return result

    def get_txt_by_xpath(self, xpath, element=None):
        value = ''
        for i in range(self.max_tries):
            value = self.cd.execute_script(f"""
           //console.log(arguments);
            parent = document;
            if(arguments[1]){{
                parent = arguments[1];
            }}
            //console.log(parent, 'parent');
            node = document.evaluate(arguments[0], parent, null,
            XPathResult.FIRST_ORDERED_NODE_TYPE, null ).singleNodeValue;
            //console.log(node, arguments);
            if(node){{
             node.scrollIntoView({{
             behavior: 'auto',
             block: 'center',
             inline: 'center'
             }});
             return node.innerText;
            }}
            return '';
            """, xpath, element)
            if value == '':
                time.sleep(self.wait_time)
        # logging.debug(["get_txt_by_xpath", value])
        return value.strip()

    def get_e_by_xpath(self, xpath, element=None):
        e = None
        for i in range(self.max_tries):
            e = self.cd.execute_script(f"""
           //console.log(arguments);
            parent = document;
            if(arguments[1]){{
                parent = arguments[1];
            }}
            //console.log(parent, 'parent');
            node = document.evaluate(arguments[0], parent, null,
            XPathResult.FIRST_ORDERED_NODE_TYPE, null ).singleNodeValue;
            //console.log(node);
            return node != null?node: null;
            """, xpath, element)
            if e is None:
                time.sleep(self.wait_time)
        logging.debug(["get_e_by_xpath", e])
        return e

    def get_attr_by_xpath(self, xpath: str, attr: str, element=None):
        value = ''
        for i in range(self.max_tries):
            value = self.cd.execute_script(f"""
           //console.log(arguments);
            parent = document;
            if(arguments[1]){{
                parent = arguments[1];
            }}
            //console.log(parent, 'parent');
            node = document.evaluate(arguments[0], parent, null,
            XPathResult.FIRST_ORDERED_NODE_TYPE, null ).singleNodeValue;
            //console.log(node);
            if(node){{
             node.scrollIntoView({{
             behavior: 'auto',
             block: 'center',
             inline: 'center'
             }});
             return node.getAttribute(arguments[2]);
            }}
            return '';
            """, xpath, element, attr)
            if value == '':
                time.sleep(self.wait_time)
        logging.debug(["get_attr_by_xpath", value])
        return value.strip()


Scraper()
